from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl


@dataclass(frozen=True)
class Provider:
    name: str
    slug: str


@dataclass(frozen=True)
class Plan:
    provider_slug: str
    name: str
    type: str


@dataclass(frozen=True)
class Price:
    provider_slug: str
    plan_name: str
    plan_type: str
    role: str
    age_min: int
    age_max: Optional[int]
    price: float
    is_particular: bool


SKIP_SHEETS = {"GALENO PD", "VISITAR"}


def norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def to_float(v: object) -> Optional[float]:
    """Convierte celdas numéricas o texto con formato es-AR (miles con punto, decimales con coma)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace("$", "").replace(" ", "")

    # Decimal con coma final (ej. 1234,56 o 1.234,56)
    if "," in s and re.search(r",\d{1,4}\s*$", s):
        s2 = s.replace(".", "") if s.count(".") > 0 else s
        s2 = s2.replace(",", ".")
        try:
            return float(s2)
        except ValueError:
            pass

    # Miles con punto(s) sin coma decimal (ej. 136.762 o 1.023.456)
    if re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
        try:
            return float(s.replace(".", ""))
        except ValueError:
            pass

    s = s.replace(".", "").replace(",", ".") if re.search(r"\d+,\d+", s) else s
    try:
        return float(s)
    except ValueError:
        return None


RANGE_RE = re.compile(
    r"(?P<min>\d+)\s*[-a]\s*(?P<max>\d+)|(?P<minplus>\d+)\s*\+|>=\s*(?P<minge>\d+)|<=\s*(?P<maxle>\d+)",
    re.IGNORECASE,
)


def parse_age_range(label: str) -> tuple[int, Optional[int]]:
    s = label.replace("AÑOS", "Años").replace("años", "Años")
    m = RANGE_RE.search(s)
    if not m:
        raise ValueError(f"No pude parsear rango etario desde: {label!r}")
    if m.group("min") and m.group("max"):
        return int(m.group("min")), int(m.group("max"))
    if m.group("minplus"):
        return int(m.group("minplus")), None
    if m.group("minge"):
        return int(m.group("minge")), None
    if m.group("maxle"):
        return 0, int(m.group("maxle"))
    raise ValueError(f"Rango etario inesperado en: {label!r}")


def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def emit_seed_sql(
    providers: list[Provider],
    plans: list[Plan],
    prices: list[Price],
    *,
    effective_month: str,
    header_title: str,
    set_active_effective_month: bool,
) -> str:
    """Genera SQL compatible con precios mensuales (effective_month + upsert por índice único)."""
    lines: list[str] = []
    lines.append(f"-- Import generado desde Excel ({header_title})")
    lines.append(f"-- Vigencia: effective_month = {effective_month}")
    lines.append("-- Ejecutar en Supabase SQL Editor. Requiere índice único prices_unique_month")
    lines.append("--   (plan_id, role, age_min, age_max, is_particular, effective_month).")
    if not set_active_effective_month:
        lines.append("-- Luego en Admin → Precios guardá active_effective_month = " + effective_month[:7] + ".")
    lines.append("begin;")
    lines.append("")

    lines.append("-- Providers")
    lines.append("insert into public.providers (name, slug)")
    lines.append("values")
    for i, p in enumerate(providers):
        comma = "," if i < len(providers) - 1 else ""
        lines.append(f"  ({sql_str(p.name)}, {sql_str(p.slug)}){comma}")
    lines.append("on conflict (slug) do nothing;")
    lines.append("")

    lines.append("-- Plans")
    for pl in plans:
        lines.append(
            "insert into public.plans (provider_id, name, type)\n"
            f"select p.id, {sql_str(pl.name)}, {sql_str(pl.type)}\n"
            "from public.providers p\n"
            f"where p.slug = {sql_str(pl.provider_slug)}\n"
            "on conflict (provider_id, name, type) do nothing;"
        )
    lines.append("")

    lines.append("-- Prices (upsert por vigencia mensual)")
    em = sql_str(effective_month)
    for pr in prices:
        age_max_sql = "null" if pr.age_max is None else str(pr.age_max)
        lines.append(
            "insert into public.prices (plan_id, role, age_min, age_max, is_particular, price, effective_month, updated_at)\n"
            "select pl.id, "
            f"{sql_str(pr.role)}::public.price_role, {pr.age_min}, {age_max_sql}, {str(pr.is_particular).lower()}, "
            f"{pr.price:.2f}, {em}::date, now()\n"
            "from public.plans pl\n"
            "join public.providers p on p.id = pl.provider_id\n"
            f"where p.slug = {sql_str(pr.provider_slug)}\n"
            f"  and pl.name = {sql_str(pr.plan_name)}\n"
            f"  and pl.type = {sql_str(pr.plan_type)}\n"
            "on conflict (plan_id, role, age_min, age_max, is_particular, effective_month)\n"
            "do update set\n"
            "  price = excluded.price,\n"
            "  updated_at = excluded.updated_at;"
        )
    lines.append("")

    if set_active_effective_month:
        lines.append("-- Cotizador: vigencia activa global")
        lines.append(
            "insert into public.app_settings (key, value_text, updated_at)\n"
            f"values ('active_effective_month', {em}, now())\n"
            "on conflict (key) do update set\n"
            "  value_text = excluded.value_text,\n"
            "  updated_at = excluded.updated_at;"
        )
        lines.append("")

    lines.append("commit;")
    lines.append("")
    return "\n".join(lines)


def sql_str(s: str) -> str:
    return "'" + s.replace("'", "''") + "'"


def parse_omint(ws) -> tuple[list[Plan], list[Price]]:
    provider_slug = "omint"
    plans: list[Plan] = []
    prices: list[Price] = []

    # Bloques: fila con "PLAN xxxx" (suele estar en col2) y headers de rangos (0-25..60+)
    for r in range(1, ws.max_row + 1):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 24) + 1)]

        plan_cell = next((v for v in row_vals[:10] if v.upper().startswith("PLAN ")), "")
        if plan_cell == "":
            continue

        # ubicar el inicio de headers etarios: "0-25" / "0-25 Años"
        start_idx = None
        for idx, v in enumerate(row_vals, 1):
            if v.startswith("0-25"):
                start_idx = idx
                break
        if start_idx is None:
            continue

        headers = [norm(ws.cell(r, c).value) for c in range(start_idx, start_idx + 5)]
        if len(headers) != 5:
            continue
        if headers[0].startswith("0-25") and headers[-1].startswith("60"):
            plan_name = plan_cell.strip()
            plan_type = "OMINT"
            plans.append(Plan(provider_slug=provider_slug, name=plan_name, type=plan_type))

            # categorías en las 3 filas siguientes (adulto, hijo1, hijo2)
            for rr in range(r + 1, min(r + 6, ws.max_row + 1)):
                # el rótulo suele estar en col2, pero toleramos que esté en col1..col6
                label = norm(ws.cell(rr, 2).value)
                if label == "":
                    label = next((norm(ws.cell(rr, c).value) for c in range(1, 7) if norm(ws.cell(rr, c).value) != ""), "")
                if label == "":
                    continue
                if "Adulto" in label or "Cony" in label or "Fam" in label:
                    role = "adulto_conyugue"
                elif "Hijo 1" in label:
                    role = "hijo_1_menor"
                elif "Hijo 2" in label:
                    role = "hijo_2_mas_menores"
                else:
                    continue

                for c, h in zip(range(start_idx, start_idx + 5), headers):
                    v = to_float(ws.cell(rr, c).value)
                    if v is None or v <= 0:
                        continue
                    age_min, age_max = parse_age_range(h.replace("o +", "+"))
                    prices.append(
                        Price(
                            provider_slug=provider_slug,
                            plan_name=plan_name,
                            plan_type=plan_type,
                            role=role,
                            age_min=age_min,
                            age_max=age_max,
                            price=v,
                            is_particular=False,
                        )
                    )

    return plans, prices


def parse_swiss(ws, plan_names: list[str], provider_slug: str = "swiss-medical") -> tuple[list[Plan], list[Price]]:
    plans: list[Plan] = []
    prices: list[Price] = []

    aliases = {
        "SBO4": "SB04",
        "SB04": "SB04",
        "SBO2": "SB02",
        "SB02": "SB02",
    }

    # Buscar fila de headers con plan names
    header_row = None
    candidates: list[int] = []
    for r in range(1, ws.max_row + 1):
        row_raw = [norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 20) + 1)]
        row = [aliases.get(v, v) for v in row_raw]
        matches = sum(1 for x in plan_names if x in row)
        if matches >= 1:
            candidates.append(r)
    # Preferimos el header más cercano a los datos (más abajo), porque suele repetir los planes
    # en el bloque "vigente" con los valores correctos.
    if candidates:
        header_row = max(candidates)
    if header_row is None:
        return plans, prices

    # Map plan -> col
    plan_cols: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v_raw = norm(ws.cell(header_row, c).value)
        v = aliases.get(v_raw, v_raw)
        if v in plan_names:
            plan_cols[v] = c

    for pn in plan_names:
        if pn in plan_cols:
            plans.append(Plan(provider_slug=provider_slug, name=pn, type="SWISS"))

    # filas de precios: menor/igual 65 y mayor 65
    for r in range(header_row + 1, min(header_row + 20, ws.max_row + 1)):
        label = norm(ws.cell(r, 1).value)
        if label == "":
            label = next(
                (norm(ws.cell(r, c).value) for c in range(2, 6) if norm(ws.cell(r, c).value) != ""),
                "",
            )
        if label == "":
            continue
        label_u = label.upper()
        if "MENOR" in label_u or "IGUAL" in label_u:
            age_min, age_max = 0, 65
        elif "MAYOR" in label_u:
            age_min, age_max = 66, None
        elif "CAPIT" in label_u:
            # Hojas Swiss tipo "SW SB02": fila "Por cápita" = misma cuota base 0-65 para todos los roles.
            age_min, age_max = 0, 65
        else:
            continue

        for pn, c in plan_cols.items():
            v = to_float(ws.cell(r, c).value)
            if v is None or v <= 0:
                continue
            # SwissStrategy suma lineal; en el Excel se ve "por cápita": lo aplicamos a todos los roles típicos
            for role in ("individual", "conyuge", "primer_hijo", "familiar_cargo"):
                prices.append(
                    Price(
                        provider_slug=provider_slug,
                        plan_name=pn,
                        plan_type="SWISS",
                        role=role,
                        age_min=age_min,
                        age_max=age_max,
                        price=v,
                        is_particular=False,
                    )
                )

    return plans, prices


def _medife_bronze_triplet_starts(ws, header_r: int) -> list[int]:
    """Columnas donde hay encabezado BRONCE/PLATA/ORO contiguo en la fila de cabecera."""
    out: list[int] = []
    for c in range(1, max(1, ws.max_column - 2) + 1):
        if norm(ws.cell(header_r, c).value).upper() != "BRONCE":
            continue
        if norm(ws.cell(header_r, c + 1).value).upper() != "PLATA":
            continue
        if norm(ws.cell(header_r, c + 2).value).upper() != "ORO":
            continue
        out.append(c)
    return out


def parse_medife(ws) -> tuple[list[Plan], list[Price]]:
    provider_slug = "medife"
    plans = [
        Plan(provider_slug=provider_slug, name="BRONCE", type="OPCION MEDIFE"),
        Plan(provider_slug=provider_slug, name="PLATA", type="OPCION MEDIFE"),
        Plan(provider_slug=provider_slug, name="ORO", type="OPCION MEDIFE"),
    ]

    # Fila de encabezados BRONCE/PLATA/ORO: en algunos libros está en 16, en otros (p. ej. Mayo 2026) en 17+.
    header_r: Optional[int] = None
    for r in range(8, min(ws.max_row, 80) + 1):
        has_bronce = any(
            norm(ws.cell(r, c).value).upper() == "BRONCE" for c in range(1, min(ws.max_column, 60) + 1)
        )
        if not has_bronce:
            continue
        nxt = norm(ws.cell(r + 1, 2).value).upper()
        if "INDIV" in nxt or "MATRIMONIO" in nxt:
            header_r = r
            break
    if header_r is None:
        return plans, []

    triplets = _medife_bronze_triplet_starts(ws, header_r)
    if not triplets:
        return plans, []

    # Si hay varias ternas (escenarios), usamos la que tenga mayor BRONCE en la primera fila INDIV
    # (suele ser el tarifario vigente a la derecha; si el libro trae solo una terna, no cambia el resultado).
    first_indiv_r: Optional[int] = None
    for r in range(header_r + 1, min(ws.max_row, header_r + 35) + 1):
        if norm(ws.cell(r, 2).value).upper().startswith("INDIV"):
            first_indiv_r = r
            break
    if first_indiv_r is not None:
        start = max(
            triplets,
            key=lambda c: to_float(ws.cell(first_indiv_r, c).value) or -1.0,
        )
    else:
        start = triplets[0]

    cols = {"BRONCE": start, "PLATA": start + 1, "ORO": start + 2}

    prices: list[Price] = []

    # Helpers de filas a roles/rangos
    def add_row(label: str, role: str, age_min: int, age_max: Optional[int], values: dict[str, float]):
        for plan_name, v in values.items():
            prices.append(
                Price(
                    provider_slug=provider_slug,
                    plan_name=plan_name,
                    plan_type="OPCION MEDIFE",
                    role=role,
                    age_min=age_min,
                    age_max=age_max,
                    price=v,
                    is_particular=False,
                )
            )

    individual_ranges: list[tuple[int, Optional[int], dict[str, float]]] = []
    for r in range(header_r + 1, min(ws.max_row, 120) + 1):
        label = norm(ws.cell(r, 2).value)
        if label == "":
            continue
        v_b = to_float(ws.cell(r, cols["BRONCE"]).value)
        v_p = to_float(ws.cell(r, cols["PLATA"]).value)
        v_o = to_float(ws.cell(r, cols["ORO"]).value)
        if v_b is None and v_p is None and v_o is None:
            continue
        values = {"BRONCE": v_b or 0.0, "PLATA": v_p or 0.0, "ORO": v_o or 0.0}

        label_u = label.upper()
        if label_u.startswith("INDIV"):
            # "INDIV(0 a 29)" / "INDIV (30 a 39)" / "INDIV (60 en +)"
            age_min, age_max = parse_age_range(label.replace("en +", "+").replace(" a ", "-"))
            add_row(label, "individual", age_min, age_max, values)
            individual_ranges.append((age_min, age_max, values))
        elif label_u.startswith("MATRIMONIO"):
            age_min, age_max = parse_age_range(label.replace("en +", "+").replace(" a ", "-"))
            add_row(label, "matrimonio", age_min, age_max, values)
        elif "1ER H" in label_u:
            add_row(label, "primer_hijo", 0, 20, values)
        elif "2DO H" in label_u:
            add_row(label, "segundo_hijo", 0, 20, values)
        elif "FAMILIAR" in label_u:
            # en el Excel aparece sin rango, lo tratamos como 0+
            add_row(label, "familiar_cargo", 0, None, values)
        else:
            continue

    # Para Medife, hijo_adulto necesita rangos para >=21: lo mapeamos a los mismos precios del adulto por rango.
    for age_min, age_max, values in individual_ranges:
        add_row("Hijo adulto (derivado de individual)", "hijo_adulto", age_min, age_max, values)

    return plans, prices


def parse_ospadep_salud(ws) -> tuple[list[Plan], list[Price]]:
    """OS 25 y OS 300 comparten filas FRANJA (columnas distintas); OS 900 es un bloque mas abajo."""
    provider_slug = "ospadep"
    plans: list[Plan] = []
    prices: list[Price] = []

    def row_has_os25_price_header(r: int) -> bool:
        v6 = norm(ws.cell(r, 6).value).lower()
        return "os 25" in norm(ws.cell(r, 2).value).lower() and "con obra social" in v6

    row_os25_header = next((r for r in range(1, min(ws.max_row, 120) + 1) if row_has_os25_price_header(r)), None)
    if row_os25_header is None:
        return plans, prices

    os300_col: Optional[int] = None
    for c in range(1, min(ws.max_column, 30) + 1):
        if norm(ws.cell(row_os25_header, c).value) == "OS 300":
            os300_col = c
            break
    if os300_col is None:
        return plans, prices
    os300_con, os300_sin = os300_col + 1, os300_col + 2

    row_os900_header = next(
        (
            r
            for r in range(row_os25_header + 1, min(ws.max_row, 200) + 1)
            if norm(ws.cell(r, 2).value) == "OS 900" and "con obra social" in norm(ws.cell(r, 6).value).lower()
        ),
        None,
    )

    plan_blocks: list[tuple[str, int, int, int, int]] = []
    r_os25_data_start = row_os25_header + 1
    r_os25_data_end = (row_os900_header - 1) if row_os900_header else ws.max_row
    plan_blocks.append(("OS 25", 6, 7, r_os25_data_start, r_os25_data_end))
    plan_blocks.append(("OS 300", os300_con, os300_sin, r_os25_data_start, r_os25_data_end))
    plans.extend(
        [
            Plan(provider_slug=provider_slug, name="OS 25", type="OSPADEP SALUD"),
            Plan(provider_slug=provider_slug, name="OS 300", type="OSPADEP SALUD"),
        ]
    )

    if row_os900_header is not None:
        r_os900_start = row_os900_header + 1
        r_os900_end = ws.max_row
        plan_blocks.append(("OS 900", 6, 7, r_os900_start, r_os900_end))
        plans.append(Plan(provider_slug=provider_slug, name="OS 900", type="OSPADEP SALUD"))

    def normalize_franja_label(label: str) -> str:
        return re.sub(r"A\s*.?\s*OS", "AÑOS", label, flags=re.IGNORECASE)

    for plan_name, con_col, sin_col, r_lo, r_hi in plan_blocks:
        for r in range(r_lo, r_hi + 1):
            label = norm(ws.cell(r, 2).value)
            if not label.upper().startswith("FRANJA") and "HIJO" not in label.upper():
                continue

            if label.upper().startswith("FRANJA"):
                age_min, age_max = parse_age_range(normalize_franja_label(label))
                roles = ("individual", "conyuge")
            else:
                age_min, age_max = 0, None
                roles = ("primer_hijo",)

            con_v = to_float(ws.cell(r, con_col).value)
            sin_v = to_float(ws.cell(r, sin_col).value)
            if con_v is not None and con_v > 0:
                for role in roles:
                    prices.append(
                        Price(
                            provider_slug=provider_slug,
                            plan_name=plan_name,
                            plan_type="OSPADEP SALUD",
                            role=role,
                            age_min=age_min,
                            age_max=age_max,
                            price=con_v,
                            is_particular=False,
                        )
                    )
            if sin_v is not None and sin_v > 0:
                for role in roles:
                    prices.append(
                        Price(
                            provider_slug=provider_slug,
                            plan_name=plan_name,
                            plan_type="OSPADEP SALUD",
                            role=role,
                            age_min=age_min,
                            age_max=age_max,
                            price=sin_v,
                            is_particular=True,
                        )
                    )

    return plans, prices


def load_workbook_from_excel(excel_path: Path):
    return openpyxl.load_workbook(excel_path, data_only=True)


def build_seed_from_workbook(
    wb,
    *,
    header_title: str,
    effective_month: str,
    set_active_effective_month: bool,
) -> str:
    providers = [
        Provider(name="OSPADEP", slug="ospadep"),
        Provider(name="Medife", slug="medife"),
        Provider(name="Omint", slug="omint"),
        Provider(name="Swiss Medical", slug="swiss-medical"),
    ]

    plans: list[Plan] = []
    prices: list[Price] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        if sheet_name == "OMINT":
            pl, pr = parse_omint(ws)
        elif sheet_name == "MEDIFE":
            pl, pr = parse_medife(ws)
        elif sheet_name == "OSPADEP SALUD":
            pl, pr = parse_ospadep_salud(ws)
        elif sheet_name.startswith("SW "):
            if sheet_name == "SW NUBIAL":
                pl, pr = parse_swiss(ws, plan_names=["MS"])
            elif sheet_name == "SW SB02":
                pl, pr = parse_swiss(ws, plan_names=["SB02"])
            else:
                pl, pr = parse_swiss(ws, plan_names=["PO62", "PO64", "SB04"])
        else:
            pl, pr = [], []

        plans.extend(pl)
        prices.extend(pr)

    plans_unique: dict[tuple[str, str, str], Plan] = {}
    for pl in plans:
        plans_unique[(pl.provider_slug, pl.name, pl.type)] = pl
    plans = list(plans_unique.values())
    plans.sort(key=lambda x: (x.provider_slug, x.type, x.name))

    prices_unique: dict[tuple, Price] = {}
    for pr in prices:
        key = (
            pr.provider_slug,
            pr.plan_name,
            pr.plan_type,
            pr.role,
            pr.age_min,
            pr.age_max,
            pr.is_particular,
        )
        prices_unique[key] = pr
    prices = list(prices_unique.values())
    prices.sort(
        key=lambda x: (
            x.provider_slug,
            x.plan_type,
            x.plan_name,
            x.role,
            x.is_particular,
            x.age_min,
            999 if x.age_max is None else x.age_max,
        )
    )

    return emit_seed_sql(
        providers=providers,
        plans=plans,
        prices=prices,
        effective_month=effective_month,
        header_title=header_title,
        set_active_effective_month=set_active_effective_month,
    )


def main() -> int:
    repo_root = Path(__file__).resolve().parents[2]
    default_excel = repo_root / "PLANES A LA VENTA - MAYO 2026.xlsx"
    default_out = Path(__file__).resolve().parents[1] / "docs" / "import-prices-from-excel.sql"

    parser = argparse.ArgumentParser(
        description="Genera SQL de importación de precios desde el Excel de planes (Supabase, vigencia mensual).",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=default_excel,
        help=f"Ruta al .xlsx (default: {default_excel})",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=default_out,
        help=f"Archivo .sql de salida (default: {default_out})",
    )
    parser.add_argument(
        "--effective-month",
        type=str,
        default="2026-05-01",
        help="Primer día del mes de vigencia en formato YYYY-MM-DD (default: 2026-05-01)",
    )
    parser.add_argument(
        "--title",
        type=str,
        default="MAYO 2026",
        help="Texto descriptivo para el encabezado del SQL",
    )
    parser.add_argument(
        "--set-active-effective-month",
        action="store_true",
        help="Incluye upsert de app_settings.active_effective_month al mismo mes",
    )
    args = parser.parse_args()

    excel_path = args.excel.resolve()
    if not excel_path.is_file():
        print(f"Error: no existe el archivo {excel_path}")
        return 1

    wb = load_workbook_from_excel(excel_path)
    seed = build_seed_from_workbook(
        wb,
        header_title=args.title,
        effective_month=args.effective_month,
        set_active_effective_month=args.set_active_effective_month,
    )

    out_path = args.out.resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(seed, encoding="utf-8")
    n_prices = seed.count("insert into public.prices (plan_id, role")
    n_plans = seed.count("insert into public.plans (provider_id")
    print(f"OK: SQL generado en {out_path}")
    print(f"Inserts planes: {n_plans} | Upserts precios: {n_prices}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

